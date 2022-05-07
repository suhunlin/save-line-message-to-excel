import os
from openpyxl import Workbook

def read_file(filename):
    messages = []
    try:
        with open(filename,'r',encoding='utf-8-sig') as file:
            for message in file:
                messages.append(message.strip())
    except Exception as error_message:
        print(error_message)

    return messages

def handle_message(messages):
    str = ''
    correct_message = []
    if not messages:
        print('無法讀取訊息資訊，請檢查訊息!!!')
    else:
        for message in messages:
            split_message_by_blank = message.split()
            if len(split_message_by_blank) > 3:
                for item in range(len(split_message_by_blank) - 2):
                    str += split_message_by_blank[2]
                    split_message_by_blank.pop(2)
                split_message_by_blank.append(str)
                correct_message.append(split_message_by_blank)
                str = ''
            else:
                correct_message.append(split_message_by_blank)

    return correct_message

def store_file_to_excel(save_filename,messages):
    wb = Workbook()
    ws = wb.active
    index = 2
    ws['A1'] = '時間'
    ws['B1'] = '姓名'
    ws['C1'] = '訊息內容'
    for message in messages:
        ws['A' + str(index)] = message[0]
        ws['B' + str(index)] = message[1]
        ws['C' + str(index)] = message[2]
        index += 1

    wb.save(save_filename)

def main():
    load_filename = 'input.txt'
    store_filename = 'message.xlsx'
    if os.path.exists(load_filename):
        messages = read_file(load_filename)
        print('檔案路徑讀取成功!!!')
        messages = handle_message(messages)
        store_file_to_excel(store_filename,messages)
    else:
        messages = []
        print('請輸入正確的檔案名稱及路徑!!!')

if __name__ == '__main__':
    main()